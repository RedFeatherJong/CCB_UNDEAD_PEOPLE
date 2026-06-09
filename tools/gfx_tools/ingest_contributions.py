#!/usr/bin/env python3
"""
读取贴图美术师填好的 Excel（内嵌 png），校验合格后自动归置贴图。

流程：
  1. 解压 xlsx，提取每个工作表内嵌的图片及其锚定行
  2. 对每个"带图片"的行做合格校验（见下）
  3. 仅当全部带图行都合格时，才执行归置：
       - png 移动到 pngs_<尺寸>/<类别子目录>/<完整贴图ID>.png
       - 生成同名 .json（CDDA 散图格式）
  4. 在脚本同目录输出"检测/贡献表"xlsx，逐行记录通过/原因/归置路径

合格判定（全部满足才通过）：
  - 完成状态 == "已完成"
  - 贡献者 非空
  - png 像素尺寸在 tile_info.json 定义的尺寸集合内
  （内嵌图片本身无文件名，落地时按该行「完整贴图ID」自动命名，无需名称比对）

只要有一行不合格，整体不归置（只出检测表）。可用 --dry-run 强制只检测。

用法：
  python3 tools/gfx_tools/ingest_contributions.py <填好的.xlsx>
  python3 tools/gfx_tools/ingest_contributions.py <填好的.xlsx> --dry-run
  python3 tools/gfx_tools/ingest_contributions.py <填好的.xlsx> --repo . --force
"""
import argparse
import json
import os
import re
import shutil
import struct
import sys
import tempfile
import zipfile
from pathlib import Path


# ---- 工作表类别 -> 子目录候选（按优先级，复用已存在的） ----
# 注意命名不统一：ChibiNormal 用 monster/item 单数，normal 用 monsters/items 复数
CATEGORY_SUBDIRS = {
    '主贴图-物品': ['items', 'item'],
    '主贴图-怪物': ['monsters', 'monster'],
    '尸体': ['monsters', 'monster'],
    '主贴图-地形': ['terrain'],
    '主贴图-家具': ['furniture'],
    '主贴图-载具部件': ['vehicle'],
    '主贴图-陷阱场效': ['traps', 'fields', 'trap', 'field'],
    'overmap': ['overmap'],
    '地图事件': ['overmap/extra'],
    '穿戴-中性': ['overlay/worn'],
    '穿戴-男': ['overlay/worn'],
    '穿戴-女': ['overlay/worn'],
    '手持': ['overlay/wielded'],
    '变异生化叠加': ['overlay/mutations'],
}

# 角色/怪物类走 Chibi 图集，其余走世界物体（normal）图集
CHIBI_CATEGORIES = {'主贴图-怪物', '尸体'}

# 列名（与 tileset_workbook.py 输出一致）
COL = {'类型': 0, 'ID': 1, '名称': 2, '前缀': 3, '完整贴图ID': 4,
       '当前贴图': 5, '完成状态': 6, '提交': 7, '贡献者': 8, '备注': 9}


def png_size(path):
    """读取 PNG 宽高（不依赖 PIL）。返回 (w,h) 或 None。"""
    try:
        with open(path, 'rb') as f:
            head = f.read(24)
        if head[:8] != b'\x89PNG\r\n\x1a\n' or head[12:16] != b'IHDR':
            return None
        w, h = struct.unpack('>II', head[16:24])
        return (w, h)
    except Exception:
        return None


def load_defined_sizes(repo):
    """从 tile_info.json 读出所有已定义的 (w,h) 尺寸集合。"""
    sizes = set()
    info_path = Path(repo) / 'tile_info.json'
    try:
        info = json.load(open(info_path, encoding='utf-8'))
    except Exception:
        return sizes
    base_w = info[0].get('width', 32) if info else 32
    base_h = info[0].get('height', 32) if info else 32
    for sheet in (info[1:] if isinstance(info, list) else []):
        for key, spec in sheet.items():
            if not key.endswith('.png') or not isinstance(spec, dict):
                continue
            w = spec.get('sprite_width', base_w)
            h = spec.get('sprite_height', base_h)
            sizes.add((w, h))
    sizes.add((base_w, base_h))
    return sizes


def find_size_dirs(repo, w, h):
    """返回该尺寸对应的所有 pngs_<root>_<w>x<h> 目录（Path 列表）。"""
    suffix = f'_{w}x{h}'
    out = []
    for d in sorted(Path(repo).glob(f'pngs_*{suffix}')):
        if d.is_dir():
            out.append(d)
    return out


def pick_size_dir(repo, w, h, category):
    """按类别在同尺寸目录里选父目录：角色/怪物走 Chibi，其余走非 Chibi。"""
    dirs = find_size_dirs(repo, w, h)
    if not dirs:
        return None
    chibi = [d for d in dirs if 'Chibi' in d.name]
    plain = [d for d in dirs
             if 'Chibi' not in d.name
             and not any(x in d.name for x in ('filler', 'incomplete',
                                               'dcss', 'fallback',
                                               'human_body'))]
    if category in CHIBI_CATEGORIES:
        return (chibi or plain or dirs)[0]
    return (plain or chibi or dirs)[0]


def resolve_subdir(parent, category):
    """在父目录下选子目录：优先复用已存在的候选，否则用第一个候选（可含 /）。"""
    candidates = CATEGORY_SUBDIRS.get(category, ['other'])
    for c in candidates:
        if (parent / c).is_dir():
            return parent / c
    return parent / candidates[0]


def parse_rels(text):
    """解析 .rels，返回 {rId: Target}。属性顺序不固定，逐个 Relationship 分别抓。"""
    out = {}
    for m in re.finditer(r'<Relationship\b[^>]*?/?>', text):
        tag = m.group(0)
        rid = re.search(r'Id="([^"]+)"', tag)
        tgt = re.search(r'Target="([^"]+)"', tag)
        if rid and tgt:
            out[rid.group(1)] = tgt.group(1)
    return out


def extract_images(xlsx_path):
    """解压 xlsx，返回 {sheet_name: {row(1基): image_abs_path}}。

    依赖 openpyxl 拿到 sheet 顺序与名称，手工解析 drawing 锚点。
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    sheet_names = wb.sheetnames

    tmp = tempfile.mkdtemp(prefix='ingest_')
    with zipfile.ZipFile(xlsx_path) as z:
        z.extractall(tmp)

    result = {sn: {} for sn in sheet_names}

    # workbook.xml: rId -> sheet name
    wb_xml = Path(tmp, 'xl', 'workbook.xml').read_text(encoding='utf-8')
    rid_to_name = {}
    for m in re.finditer(r'<sheet\b[^>]*?/?>', wb_xml):
        tag = m.group(0)
        nm = re.search(r'name="([^"]+)"', tag)
        rid = re.search(r'r:id="(rId\d+)"', tag)
        if nm and rid:
            rid_to_name[rid.group(1)] = nm.group(1)
    # workbook.xml.rels: rId -> sheetN.xml（属性顺序不固定，分别抓取）
    wbrels = Path(tmp, 'xl', '_rels', 'workbook.xml.rels').read_text(encoding='utf-8')
    rid_to_target = parse_rels(wbrels)
    # sheetN.xml -> sheet name
    file_to_name = {}
    for rid, name in rid_to_name.items():
        tgt = rid_to_target.get(rid, '')
        base = os.path.basename(tgt)
        if base:
            file_to_name[base] = name

    # 每个 worksheet 找它的 drawing，再解析锚点
    ws_dir = Path(tmp, 'xl', 'worksheets')
    for sheet_file in ws_dir.glob('sheet*.xml'):
        sname = file_to_name.get(sheet_file.name)
        if not sname:
            continue
        rels = ws_dir / '_rels' / (sheet_file.name + '.rels')
        if not rels.exists():
            continue
        rtext = rels.read_text(encoding='utf-8')
        dm = re.search(r'drawing(\d+)\.xml', rtext)
        if not dm:
            continue
        draw_file = Path(tmp, 'xl', 'drawings', f'drawing{dm.group(1)}.xml')
        draw_rels = Path(tmp, 'xl', 'drawings', '_rels',
                         f'drawing{dm.group(1)}.xml.rels')
        if not draw_file.exists() or not draw_rels.exists():
            continue
        dr = draw_rels.read_text(encoding='utf-8')
        embed_to_media = {rid: os.path.basename(tgt)
                          for rid, tgt in parse_rels(dr).items()}
        dtext = draw_file.read_text(encoding='utf-8')
        # 每个 anchor：from row + embed rId
        # 命名空间前缀可选：WPS/Excel 用 <xdr:from><xdr:row>，openpyxl 用 <from><row>
        for am in re.finditer(
                r'<(?:\w+:)?from>.*?<(?:\w+:)?row>(\d+)</(?:\w+:)?row>.*?'
                r'embed="(rId\d+)"', dtext, re.S):
            row0 = int(am.group(1))
            media = embed_to_media.get(am.group(2))
            if not media:
                continue
            media_path = Path(tmp, 'xl', 'drawings', 'media', media)
            if not media_path.exists():
                media_path = Path(tmp, 'xl', 'media', media)
            if media_path.exists():
                result[sname][row0 + 1] = str(media_path)  # 1基行号

    return result, tmp


def validate_row(ws, row, sheet_name, img_path, defined_sizes):
    """返回 (ok: bool, reason: str, info: dict)。"""
    def cell(name):
        v = ws.cell(row, COL[name] + 1).value
        return v.strip() if isinstance(v, str) else v

    full_id = cell('完整贴图ID') or ''
    status = cell('完成状态') or ''
    contributor = cell('贡献者')
    reasons = []

    if status != '已完成':
        reasons.append(f'完成状态为「{status or "空"}」，需为「已完成」')
    if not contributor:
        reasons.append('缺贡献者')

    # png 尺寸
    size = png_size(img_path)
    if size is None:
        reasons.append('图片无法读取或非 PNG')
    elif size not in defined_sizes:
        reasons.append(f'尺寸 {size[0]}x{size[1]} 不在已定义尺寸内')

    info = {'full_id': full_id, 'status': status,
            'contributor': contributor or '', 'size': size}
    return (len(reasons) == 0, '；'.join(reasons), info)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('xlsx', help='贴图美术师填好的 Excel')
    ap.add_argument('--repo', default='.', help='贴图源仓库根目录（含 tile_info.json）')
    ap.add_argument('--report', default='归置检测表.xlsx', help='输出检测表路径')
    ap.add_argument('--dry-run', action='store_true', help='只检测不落地')
    ap.add_argument('--force', action='store_true', help='允许覆盖已存在的同名 png')
    args = ap.parse_args()

    from openpyxl import load_workbook

    repo = Path(args.repo).resolve()
    if not (repo / 'tile_info.json').exists():
        sys.exit(f'错误：{repo} 下无 tile_info.json，--repo 需指向贴图源仓库根目录')

    defined_sizes = load_defined_sizes(repo)
    if not defined_sizes:
        sys.exit('错误：未能从 tile_info.json 读到任何已定义尺寸')

    images, tmp = extract_images(args.xlsx)
    wb = load_workbook(args.xlsx)

    # 收集所有带图行 + 校验
    records = []  # dict per image row
    for sname, rowmap in images.items():
        if not rowmap:
            continue
        ws = wb[sname]
        for row, img_path in sorted(rowmap.items()):
            ok, reason, info = validate_row(ws, row, sname, img_path, defined_sizes)
            png_name = info['full_id'] + '.png' if info['full_id'] else ''
            records.append({
                'sheet': sname, 'row': row, 'img': img_path,
                'full_id': info['full_id'], 'png_name': png_name,
                'status': info['status'], 'contributor': info['contributor'],
                'size': info['size'], 'ok': ok, 'reason': reason,
                'dest': '',
            })

    if not records:
        print('未在表格中找到任何内嵌图片，无事可做。')
        shutil.rmtree(tmp, ignore_errors=True)
        write_report(args.report, records)
        return

    all_ok = all(r['ok'] for r in records)
    n_ok = sum(1 for r in records if r['ok'])
    print(f'带图行 {len(records)} 个，合格 {n_ok}，不合格 {len(records) - n_ok}')

    do_place = all_ok and not args.dry_run
    if not all_ok:
        print('存在不合格行，整体不归置（仅输出检测表）。')
    elif args.dry_run:
        print('--dry-run：仅检测，不落地。')

    if do_place:
        for r in records:
            place_one(r, repo, args.force)

    write_report(args.report, records)
    shutil.rmtree(tmp, ignore_errors=True)

    print(f'\n检测表已写入：{args.report}')
    if do_place:
        print(f'已归置 {sum(1 for r in records if r["dest"])} 张贴图。')


def place_one(rec, repo, force):
    """归置单张 png + 生成 json。结果写回 rec['dest'] / rec['reason']。"""
    size = rec['size']
    parent = pick_size_dir(repo, size[0], size[1], rec['sheet'])
    if parent is None:
        rec['ok'] = False
        rec['reason'] = f'无 {size[0]}x{size[1]} 对应的 pngs 目录'
        return
    dest_dir = resolve_subdir(parent, rec['sheet'])
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_png = dest_dir / rec['png_name']
    dest_json = dest_dir / (rec['full_id'] + '.json')

    if dest_png.exists() and not force:
        rec['ok'] = False
        rec['reason'] = f'目标已存在同名 png：{dest_png.relative_to(repo)}（用 --force 覆盖）'
        return

    shutil.copy2(rec['img'], dest_png)
    # CDDA 散图 json
    tile_def = [{'id': [rec['full_id']], 'fg': [rec['full_id']]}]
    dest_json.write_text(
        json.dumps(tile_def, ensure_ascii=False, indent=2) + '\n',
        encoding='utf-8')
    rec['dest'] = str(dest_png.relative_to(repo))


def write_report(path, records):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = '归置检测'
    cols = ['工作表', '行', 'ID/完整贴图ID', 'png文件名', '尺寸',
            '完成状态', '贡献者', '结果', '原因', '归置路径']
    ws.append(cols)
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='4472C4')
    okfill = PatternFill('solid', fgColor='C6EFCE')
    badfill = PatternFill('solid', fgColor='FFC7CE')
    center = Alignment(horizontal='center')
    for c in ws[1]:
        c.font, c.fill, c.alignment = hf, hfill, center
    ws.freeze_panes = 'A2'

    for r in records:
        size_str = f'{r["size"][0]}x{r["size"][1]}' if r.get('size') else '—'
        ws.append([r['sheet'], r['row'], r['full_id'], r['png_name'],
                   size_str, r['status'], r['contributor'],
                   '通过' if r['ok'] else '不通过',
                   r['reason'], r['dest']])
        ws.cell(ws.max_row, 8).fill = okfill if r['ok'] else badfill

    for col, w in {'A': 16, 'B': 6, 'C': 30, 'D': 30, 'E': 8,
                   'F': 10, 'G': 12, 'H': 8, 'I': 40, 'J': 40}.items():
        ws.column_dimensions[col].width = w
    wb.save(path)


if __name__ == '__main__':
    main()
